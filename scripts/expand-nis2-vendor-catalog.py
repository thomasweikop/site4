#!/usr/bin/env python3

from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "src/data/nis2_vendor_area_matrix_real_v4.csv"
JSON_PATH = ROOT / "src/data/nis2_vendor_area_matrix_real_v4.json"
XLSX_PATH = ROOT / "exports/complycheck_nis2_vendor_catalog_v5.xlsx"
DOWNLOAD_XLSX_PATH = (
    Path("/Users/thomasweikop/Downloads") / "complycheck_nis2_vendor_catalog_v5.xlsx"
)

QUALIFICATION_NOTE = (
    "Profilen er placeret på baggrund af offentlig markedspositionering, hjemmeside, "
    "beskrevne ydelser og relevans for NIS2-relateret rådgivning eller implementering."
)

DEFAULT_ROLE_BY_TYPE = {
    "Legal": (
        "Brug denne profil når scope, kontrakter, ledelsesansvar, myndighedsrapportering "
        "eller fortolkning af NIS2, CER eller CRA er den primære udfordring."
    ),
    "GRC": (
        "Brug denne profil når politikker, ISMS, dokumentation, risikoregister, "
        "leverandørstyring eller ledelsesrapportering er de største mangler."
    ),
    "Technical": (
        "Brug denne profil når virksomheden skal implementere tekniske kontroller, "
        "styrke arkitektur, adgangsstyring, hardening, backup eller netværkssikkerhed."
    ),
    "SOC": (
        "Brug denne profil når logging, monitorering, detektion, MDR eller incident "
        "response skal professionaliseres eller drives mere kontinuerligt."
    ),
    "Audit": (
        "Brug denne profil når virksomheden har brug for uafhængig validering, review, "
        "assurance, certificeringsforløb eller dokumentation til ledelse, kunder og partnere."
    ),
}

TARGET_COUNTS = {
    "Legal": 42,
    "GRC": 45,
    "Technical": 53,
    "SOC": 43,
    "Audit": 41,
}

TEMPLATE_BY_TYPE = {
    "Legal": "Horten",
    "GRC": "NorthGRC",
    "Technical": "Atea Danmark",
    "SOC": "CapMon",
    "Audit": "Bureau Veritas Denmark",
}

TYPE_SORT_ORDER = ["Legal", "GRC", "Technical", "SOC", "Audit"]

ADDITIONS = {
    "Legal": [
        {
            "name": "DELACOUR",
            "website": "https://www.delacour.dk",
            "best_for": "Erhvervsjuridisk NIS2-rådgivning med fokus på kontrakter, ansvar og regulerede virksomheder.",
            "score": 74,
        },
        {
            "name": "Holst, Advokater",
            "website": "https://www.holst-law.com",
            "best_for": "Forretningsjuridisk rådgivning om it, databeskyttelse og cybersikkerhedsrelaterede krav.",
            "score": 73,
        },
        {
            "name": "HjulmandKaptain",
            "website": "https://www.hjulmandkaptain.dk",
            "best_for": "Juridisk rådgivning om governance, ledelsesansvar og kontraktuelle sikkerhedskrav.",
            "score": 72,
        },
        {
            "name": "TVC Advokatfirma",
            "website": "https://www.tvc.dk",
            "best_for": "Erhvervs- og regulatorisk rådgivning til virksomheder med behov for klarere compliance-rammer.",
            "score": 71,
        },
        {
            "name": "Roesgaard Advokater",
            "website": "https://roesgaard.dk",
            "best_for": "Juridisk rådgivning om databeskyttelse, kontrakter og regulatoriske forpligtelser.",
            "score": 71,
        },
        {
            "name": "White & Case",
            "website": "https://www.whitecase.com",
            "best_for": "International juridisk rådgivning om cybersikkerhed, regulering og komplekse leverandørforhold.",
            "size_fit": "Mid-market, enterprise",
            "score": 76,
        },
        {
            "name": "Baker McKenzie",
            "website": "https://www.bakermckenzie.com",
            "best_for": "Tværnational juridisk rådgivning om NIS2, privatliv, datatransfers og regulatoriske programmer.",
            "score": 76,
        },
        {
            "name": "CMS",
            "website": "https://cms.law",
            "best_for": "EU-retlig og kommerciel rådgivning om cybersikkerhed, kontrakter og digital regulering.",
            "score": 74,
        },
        {
            "name": "Hogan Lovells",
            "website": "https://www.hoganlovells.com",
            "best_for": "International rådgivning om cyberincidenter, data, kontrakter og regulatoriske forløb.",
            "score": 76,
        },
        {
            "name": "Norton Rose Fulbright",
            "website": "https://www.nortonrosefulbright.com",
            "best_for": "Regulatorisk og kontraktuel rådgivning til større organisationer i kritiske sektorer.",
            "score": 75,
        },
        {
            "name": "Pinsent Masons",
            "website": "https://www.pinsentmasons.com",
            "best_for": "Teknologi- og regulatorisk rådgivning om cybersikkerhed, digitale kontrakter og leverandørstyring.",
            "score": 75,
        },
        {
            "name": "A&O Shearman",
            "website": "https://www.aoshearman.com",
            "best_for": "Kompleks juridisk rådgivning om cyberrisiko, ledelsesansvar og grænseoverskridende compliance.",
            "score": 76,
        },
        {
            "name": "Fieldfisher",
            "website": "https://www.fieldfisher.com",
            "best_for": "Juridisk rådgivning i krydsfeltet mellem teknologi, databeskyttelse og regulatoriske krav.",
            "score": 74,
        },
        {
            "name": "Osborne Clarke",
            "website": "https://www.osborneclarke.com",
            "best_for": "Rådgivning om digital regulering, it-kontrakter, databeskyttelse og cybersikkerhed.",
            "score": 74,
        },
        {
            "name": "Simmons & Simmons",
            "website": "https://www.simmons-simmons.com",
            "best_for": "Regulatorisk rådgivning om teknologi, databeskyttelse, cyberhændelser og kontraktstyring.",
            "score": 74,
        },
        {
            "name": "Eversheds Sutherland",
            "website": "https://www.eversheds-sutherland.com",
            "best_for": "International erhvervsjuridisk rådgivning om cyber, databeskyttelse og sektorspecifik regulering.",
            "score": 74,
        },
        {
            "name": "Addleshaw Goddard",
            "website": "https://www.addleshawgoddard.com",
            "best_for": "Juridisk rådgivning om outsourcing, kontrakter, compliance-programmer og digital regulering.",
            "score": 73,
        },
    ],
    "GRC": [
        {
            "name": "Implement Consulting Group",
            "website": "https://implementconsultinggroup.com",
            "best_for": "Governance-, transformations- og dokumentationsspor med fokus på ledelsesforankring og prioritering.",
            "score": 78,
        },
        {
            "name": "PA Consulting",
            "website": "https://www.paconsulting.com",
            "best_for": "Strategisk cyber- og risikorådgivning med fokus på operating model, governance og regulerede miljøer.",
            "score": 78,
        },
        {
            "name": "BearingPoint",
            "website": "https://www.bearingpoint.com",
            "best_for": "GRC, risikostyring, modenhedsvurderinger og dokumentationsforløb til større virksomheder.",
            "score": 77,
        },
        {
            "name": "Devoteam Denmark",
            "website": "https://www.devoteam.com",
            "best_for": "GRC- og transformationsrådgivning med kobling mellem governance, cloud og sikkerhedsdrift.",
            "score": 77,
        },
        {
            "name": "Accenture Denmark Cyber Strategy & GRC",
            "website": "https://www.accenture.com/dk-en",
            "best_for": "Strategi-, governance- og complianceprogrammer i større eller komplekse organisationer.",
            "score": 79,
        },
        {
            "name": "Eviden Cyber GRC Denmark",
            "website": "https://eviden.com",
            "best_for": "GRC-forløb med fokus på modenhed, dokumentation, kontrolframeworks og cyberprogramstyring.",
            "score": 78,
        },
        {
            "name": "Sopra Steria Cyber Advisory Denmark",
            "website": "https://www.soprasteria.com",
            "best_for": "Governance, operating model og dokumentationsarbejde i offentlige og regulerede miljøer.",
            "sector_fit": "Public sector, finance, healthcare, transport, critical sectors",
            "score": 77,
        },
        {
            "name": "Crayon Denmark Governance & Security",
            "website": "https://www.crayon.com",
            "best_for": "Governance- og licensnære sikkerhedsforløb omkring cloud, identitet og dokumentation.",
            "score": 75,
        },
        {
            "name": "NTT DATA Business Solutions Denmark Advisory",
            "website": "https://www.nttdata-solutions.com",
            "best_for": "Governance og compliance i it- og ERP-tunge miljøer med mange forretningskritiske processer.",
            "score": 76,
        },
        {
            "name": "Aeven Governance & Compliance",
            "website": "https://www.aeven.com",
            "best_for": "Governance, risikostyring og driftsnære complianceforløb i større it-landskaber.",
            "score": 76,
        },
        {
            "name": "Omada Identity Advisory",
            "website": "https://omadaidentity.com",
            "best_for": "Governance omkring identiteter, roller, access governance og kontrollerbar dokumentation.",
            "sector_fit": "Finance, SaaS, tech and regulated services",
            "score": 79,
        },
        {
            "name": "RiskPoint Group",
            "website": "https://www.riskpoint.eu",
            "best_for": "Risikostyring, compliance og cyber governance med fokus på forsikrings- og enterprise-miljøer.",
            "score": 76,
        },
        {
            "name": "Truesec Governance, Risk & Compliance",
            "website": "https://www.truesec.com",
            "best_for": "GRC-forløb med kobling mellem governance, security operations og prioriterede forbedringsplaner.",
            "score": 78,
        },
        {
            "name": "mnemonic Advisory",
            "website": "https://www.mnemonic.io",
            "best_for": "Cyber governance, modenhedsvurderinger og driftsnære forbedringsforløb i større miljøer.",
            "score": 77,
        },
        {
            "name": "2Secure Governance & Compliance",
            "website": "https://2secure.se",
            "best_for": "Governance, risikostyring og compliance med stærk kobling til praktisk sikkerhedsdrift.",
            "score": 77,
        },
        {
            "name": "BSI Group Advisory",
            "website": "https://www.bsigroup.com",
            "best_for": "ISMS-, governance- og standardbaserede complianceforløb med fokus på dokumentation og modenhed.",
            "score": 78,
        },
        {
            "name": "LRQA Cyber & Compliance",
            "website": "https://www.lrqa.com",
            "best_for": "Complianceprogrammer, modenhedsvurderinger og styringsforløb med afsæt i standarder og krav.",
            "score": 77,
        },
        {
            "name": "OneTrust Services",
            "website": "https://www.onetrust.com",
            "best_for": "Dokumentations- og governancearbejde i krydsfeltet mellem privacy, third-party risk og compliance.",
            "sector_fit": "Cross-sector; prioritize critical sectors",
            "score": 74,
        },
        {
            "name": "Vanta Advisory Partners",
            "website": "https://www.vanta.com",
            "best_for": "Dokumentations- og kontrolprogrammer for virksomheder der vil systematisere compliance hurtigere.",
            "size_fit": "SMB, mid-market",
            "sector_fit": "SaaS, software, digital, IT",
            "score": 72,
        },
        {
            "name": "Drata Advisory Partners",
            "website": "https://drata.com",
            "best_for": "Kontrolbiblioteker, dokumentation og modenhedsstyring for digitale og hurtigt voksende virksomheder.",
            "size_fit": "SMB, mid-market",
            "sector_fit": "SaaS, software, digital, IT",
            "score": 72,
        },
    ],
    "Technical": [
        {
            "name": "Accenture Denmark Cybersecurity Engineering",
            "website": "https://www.accenture.com/dk-en",
            "best_for": "Teknisk implementering af sikkerhedskontroller i komplekse eller internationale miljøer.",
            "score": 79,
        },
        {
            "name": "Eviden Cybersecurity Services Denmark",
            "website": "https://eviden.com",
            "best_for": "Security engineering, identitet, netværk, endpoint og tekniske moderniseringsforløb.",
            "score": 78,
        },
        {
            "name": "Sopra Steria Security Engineering",
            "website": "https://www.soprasteria.com",
            "best_for": "Teknisk sikkerhedsarkitektur og implementering i offentlige og regulerede miljøer.",
            "sector_fit": "Public sector, finance, healthcare, transport, critical sectors",
            "score": 77,
        },
        {
            "name": "Fujitsu Denmark Security Services",
            "website": "https://www.fujitsu.com/dk",
            "best_for": "Tekniske sikkerhedsforbedringer, drift, infrastruktursikkerhed og recovery-forløb.",
            "score": 76,
        },
        {
            "name": "Devoteam Denmark Cloud & Cyber",
            "website": "https://www.devoteam.com",
            "best_for": "Cloud security, identitet, governance-nære kontroller og praktisk implementering.",
            "score": 76,
        },
        {
            "name": "Advania Denmark Security Services",
            "website": "https://www.advania.dk",
            "best_for": "Pragmatisk implementering af netværk, endpoint, backup og adgangskontroller.",
            "score": 76,
        },
        {
            "name": "Dustin Denmark Security & Infrastructure",
            "website": "https://www.dustin.dk",
            "best_for": "Infrastruktur- og sikkerhedsforbedringer i mid-market-miljøer med fokus på basiskontroller.",
            "size_fit": "SMB, mid-market",
            "score": 74,
        },
        {
            "name": "Crayon Denmark Security Implementation",
            "website": "https://www.crayon.com",
            "best_for": "Cloudnær sikkerhedsimplementering, identitet og styrkelse af Microsoft-baserede kontroller.",
            "score": 75,
        },
        {
            "name": "NTT DATA Business Solutions Denmark Security",
            "website": "https://www.nttdata-solutions.com",
            "best_for": "Implementering af tekniske sikkerhedskontroller i ERP- og forretningskritiske platforme.",
            "score": 76,
        },
        {
            "name": "Aeven Security Engineering",
            "website": "https://www.aeven.com",
            "best_for": "Teknisk sikkerhed, recovery, platformdrift og forbedring af centrale kontroller i enterprise-miljøer.",
            "score": 77,
        },
        {
            "name": "Cegal Danmark Security Services",
            "website": "https://www.cegal.com",
            "best_for": "Teknisk sikkerhed og robust it-drift i kritiske eller dataintensive miljøer.",
            "sector_fit": "Energy, software, digital, IT, critical sectors",
            "score": 75,
        },
        {
            "name": "Microsoft Danmark Security Advisory",
            "website": "https://www.microsoft.com/da-dk",
            "best_for": "Identitet, cloud security, logging og tekniske grundkontroller i Microsoft-tunge miljøer.",
            "sector_fit": "Cross-sector; prioritize critical sectors",
            "score": 75,
        },
        {
            "name": "HCLTech Denmark Cybersecurity Services",
            "website": "https://www.hcltech.com",
            "best_for": "Større implementeringsforløb for sikkerhedsarkitektur, endpoint, recovery og drift.",
            "score": 74,
        },
        {
            "name": "Wipro Denmark Cyber Defense Engineering",
            "website": "https://www.wipro.com",
            "best_for": "Teknisk implementering af sikkerhedskontroller i globale eller komplekse driftsmiljøer.",
            "score": 74,
        },
        {
            "name": "Tata Consultancy Services Denmark Cybersecurity",
            "website": "https://www.tcs.com",
            "best_for": "Større transformations- og implementeringsforløb med fokus på standardisering og teknisk modenhed.",
            "score": 74,
        },
        {
            "name": "Infosys Denmark Cybersecurity Services",
            "website": "https://www.infosys.com",
            "best_for": "Teknisk sikkerhedsimplementering, adgangsstyring og modernisering af kontrolmiljøer.",
            "score": 74,
        },
        {
            "name": "Capgemini Denmark Cybersecurity",
            "website": "https://www.capgemini.com",
            "best_for": "Arkitektur, sikkerhedsprogrammer og teknisk implementering i større eller regulerede virksomheder.",
            "score": 76,
        },
        {
            "name": "Sentia Denmark Cloud Security",
            "website": "https://www.sentia.com",
            "best_for": "Cloud security, platform hardening og driftsnær implementering af sikkerhedsforbedringer.",
            "score": 75,
        },
        {
            "name": "Venzo Cyber Security",
            "website": "https://venzo.com",
            "best_for": "Teknisk cybersikkerhed, rådgivning og implementering med fokus på identitet og modstandsdygtighed.",
            "score": 76,
        },
        {
            "name": "Truesec Security Engineering",
            "website": "https://www.truesec.com",
            "best_for": "Tekniske forbedringer i krydsfeltet mellem engineering, detektion og incident readiness.",
            "score": 77,
        },
        {
            "name": "mnemonic Incident & Security Engineering",
            "website": "https://www.mnemonic.io",
            "best_for": "Security engineering med stærk kobling til detektion, beredskab og tekniske forbedringer.",
            "score": 76,
        },
        {
            "name": "2Secure Security Engineering",
            "website": "https://2secure.se",
            "best_for": "Teknisk implementering og forbedring af drift, adgangskontroller og beredskab.",
            "score": 75,
        },
        {
            "name": "Palo Alto Networks Unit 42",
            "website": "https://unit42.paloaltonetworks.com",
            "best_for": "Avanceret security engineering, cloud security og incident readiness i større miljøer.",
            "score": 77,
        },
        {
            "name": "Fortinet Professional Services",
            "website": "https://www.fortinet.com",
            "best_for": "Implementering af netværkssikkerhed, segmentering, logging og driftsnære kontroller.",
            "score": 74,
        },
        {
            "name": "Check Point Professional Services",
            "website": "https://www.checkpoint.com",
            "best_for": "Netværks- og perimeterkontroller, segmentering og forbedring af tekniske basiskontroller.",
            "score": 74,
        },
        {
            "name": "Redpill Linpro Denmark",
            "website": "https://www.redpill-linpro.com",
            "best_for": "Open source-, platform- og cloudnære sikkerhedsforbedringer i tekniske miljøer.",
            "score": 73,
        },
        {
            "name": "Arrow ECS Denmark Security Solutions",
            "website": "https://www.arrow.com",
            "best_for": "Sikkerhedsarkitektur, platformvalg og implementering i samarbejde med enterprise-partnere.",
            "score": 73,
        },
        {
            "name": "Hewlett Packard Enterprise Denmark Security",
            "website": "https://www.hpe.com/dk/da",
            "best_for": "Infrastruktur-, netværks- og platformnære sikkerhedsforbedringer i større it-miljøer.",
            "score": 73,
        },
    ],
    "SOC": [
        {
            "name": "Truesec MDR",
            "website": "https://www.truesec.com",
            "best_for": "MDR, incident response og driftsnær monitorering for virksomheder med behov for hurtigere detektion.",
            "score": 84,
        },
        {
            "name": "mnemonic MDR",
            "website": "https://www.mnemonic.io",
            "best_for": "Monitorering, MDR og incident response for større eller mere trusselsudsatte miljøer.",
            "score": 83,
        },
        {
            "name": "2Secure Detection & Response",
            "website": "https://2secure.se",
            "best_for": "Detektion, incident response og operationalisering af sikkerhedsberedskab.",
            "score": 82,
        },
        {
            "name": "Sentia Denmark SOC Services",
            "website": "https://www.sentia.com",
            "best_for": "Cloudnær monitorering og sikkerhedsdrift i miljøer med fokus på kontinuerlig opfølgning.",
            "score": 80,
        },
        {
            "name": "Palo Alto Networks Unit 42 Managed Services",
            "website": "https://unit42.paloaltonetworks.com",
            "best_for": "Avanceret monitorering, incident response og specialistbistand ved komplekse hændelser.",
            "score": 83,
        },
        {
            "name": "Arctic Wolf",
            "website": "https://arcticwolf.com",
            "best_for": "MDR- og monitoreringsforløb for virksomheder der vil styrke detektion uden selv at bygge SOC.",
            "score": 81,
        },
        {
            "name": "WithSecure Detection & Response",
            "website": "https://www.withsecure.com",
            "best_for": "Detektion, endpoint-nær monitorering og hurtigere eskalation ved sikkerhedshændelser.",
            "score": 80,
        },
        {
            "name": "Darktrace Managed Detection & Response",
            "website": "https://www.darktrace.com",
            "best_for": "AI-understøttet detektion og monitorering i miljøer med behov for bred telemetry og hurtig respons.",
            "score": 79,
        },
        {
            "name": "Rapid7 Managed Detection & Response",
            "website": "https://www.rapid7.com",
            "best_for": "MDR, monitorering og incident response med fokus på at operationalisere eksisterende logkilder.",
            "score": 79,
        },
        {
            "name": "Sophos MDR",
            "website": "https://www.sophos.com",
            "best_for": "MDR og monitorering i virksomheder der vil have et standardiseret og hurtigt startbart SOC-spor.",
            "size_fit": "SMB, mid-market, enterprise",
            "score": 78,
        },
        {
            "name": "CrowdStrike Falcon Complete",
            "website": "https://www.crowdstrike.com",
            "best_for": "Endpoint-nær monitorering, detektion og respons i miljøer med fokus på moderne angrebsovervågning.",
            "score": 79,
        },
        {
            "name": "Fujitsu Denmark SOC Services",
            "website": "https://www.fujitsu.com/dk",
            "best_for": "Monitorering, sikkerhedsdrift og incident response i enterprise- og driftskritiske miljøer.",
            "score": 78,
        },
        {
            "name": "Accenture Managed Detection & Response",
            "website": "https://www.accenture.com/dk-en",
            "best_for": "Skalerbar SOC- og monitoreringskapacitet i større eller internationalt forankrede organisationer.",
            "score": 81,
        },
        {
            "name": "Eviden Managed Detection & Response",
            "website": "https://eviden.com",
            "best_for": "Monitorering, detektion og incident response med fokus på enterprise-drift og regulerede miljøer.",
            "score": 80,
        },
        {
            "name": "Sopra Steria SOC Services",
            "website": "https://www.soprasteria.com",
            "best_for": "Sikkerhedsdrift og monitorering i organisationer der ønsker kobling mellem drift, governance og SOC.",
            "score": 79,
        },
        {
            "name": "Tietoevry Managed Detection & Response",
            "website": "https://www.tietoevry.com",
            "best_for": "Nordisk sikkerhedsdrift og monitorering med fokus på stabil operation og kontinuerlig forbedring.",
            "score": 79,
        },
        {
            "name": "Microsoft Security Experts",
            "website": "https://www.microsoft.com/da-dk/security/business/services/microsoft-security-experts",
            "best_for": "Detektion og forbedring af monitorering i Microsoft-baserede miljøer med fokus på hurtig operationalisering.",
            "score": 78,
        },
        {
            "name": "NTT DATA Managed Security Services",
            "website": "https://www.nttdata.com",
            "best_for": "Monitorering, SOC og sikkerhedsdrift i større eller proceskritiske it-landskaber.",
            "score": 79,
        },
        {
            "name": "Wipro Cyber Defense Center",
            "website": "https://www.wipro.com",
            "best_for": "Monitorering, detektion og incident response for virksomheder med behov for skalerbar kapacitet.",
            "score": 78,
        },
        {
            "name": "Capgemini SOC & Monitoring",
            "website": "https://www.capgemini.com",
            "best_for": "SOC- og monitoreringskapacitet med fokus på drift, modenhed og integreret cybersikkerhedsleverance.",
            "score": 79,
        },
    ],
    "Audit": [
        {
            "name": "DEKRA Certification Denmark",
            "website": "https://www.dekra.com",
            "best_for": "Ekstern validering, certificering og modenhedsvurderinger af sikkerheds- og compliancekontroller.",
            "score": 78,
        },
        {
            "name": "Intertek Denmark",
            "website": "https://www.intertek.com",
            "best_for": "Assurance og review af processer, kontroller og standardnære krav i regulerede miljøer.",
            "score": 76,
        },
        {
            "name": "SGS Denmark",
            "website": "https://www.sgs.com",
            "best_for": "Ekstern assurance, certificering og verificering af styrings- og kontrolmiljøer.",
            "score": 76,
        },
        {
            "name": "Kiwa Denmark",
            "website": "https://www.kiwa.com/dk",
            "best_for": "Assurance- og certificeringsforløb med fokus på dokumentation, processer og efterviselig modenhed.",
            "score": 76,
        },
        {
            "name": "TUV SUD Denmark",
            "website": "https://www.tuvsud.com",
            "best_for": "Certificering, uafhængigt review og validering af processer og sikkerhedsrelaterede kontroller.",
            "score": 76,
        },
        {
            "name": "TUV Rheinland Denmark",
            "website": "https://www.tuv.com",
            "best_for": "Ekstern assurance og certificeringsnære vurderinger af styrings- og kontrolrammer.",
            "score": 76,
        },
        {
            "name": "LRQA Assurance",
            "website": "https://www.lrqa.com",
            "best_for": "Assurance, review og modenhedsvurderinger af governance, dokumentation og leverandørstyring.",
            "score": 77,
        },
        {
            "name": "BSI Group Assurance",
            "website": "https://www.bsigroup.com",
            "best_for": "Uafhængige review- og assuranceforløb baseret på standarder, ISMS og dokumenterede kontroller.",
            "score": 77,
        },
        {
            "name": "ControlCase",
            "website": "https://www.controlcase.com",
            "best_for": "Assurance og review af kontrolmiljøer, modenhed og dokumentation i internationale miljøer.",
            "score": 75,
        },
        {
            "name": "FORCE Technology",
            "website": "https://forcetechnology.com",
            "best_for": "Test-, validerings- og assuranceforløb i tekniske eller regulerede miljøer.",
            "score": 75,
        },
        {
            "name": "Nemko",
            "website": "https://www.nemko.com",
            "best_for": "Ekstern validering, test og compliancevurderinger af tekniske og procesnære kontroller.",
            "score": 74,
        },
        {
            "name": "A-LIGN",
            "website": "https://www.a-lign.com",
            "best_for": "Assurance og rapporteringsforløb med fokus på dokumentation, kontroller og modenhedsbevis.",
            "score": 74,
        },
        {
            "name": "Schellman",
            "website": "https://www.schellman.com",
            "best_for": "Uafhængige review- og assuranceforløb i organisationer med høje krav til efterviselighed.",
            "score": 74,
        },
        {
            "name": "NQA",
            "website": "https://www.nqa.com",
            "best_for": "Certificering og assurance med fokus på styringssystemer, dokumentation og opfølgning.",
            "score": 73,
        },
        {
            "name": "QMS International",
            "website": "https://www.qmsuk.com",
            "best_for": "Certificeringsnære review- og assuranceforløb for virksomheder der vil dokumentere modenhed.",
            "score": 72,
        },
        {
            "name": "IASME Consortium",
            "website": "https://iasme.co.uk",
            "best_for": "Assurance- og modenhedsvurderinger med fokus på praktisk dokumentation og basiskontroller.",
            "size_fit": "SMB, mid-market",
            "score": 71,
        },
    ],
}


def load_csv_rows() -> tuple[list[dict[str, str]], list[str]]:
    with CSV_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return list(reader), reader.fieldnames or []


def size_fit_text(size_fit: str) -> str:
    text = size_fit.lower()
    if "smb" in text and "enterprise" in text:
        return "fra mindre virksomheder til større enterprise-miljøer"
    if "smb" in text and "mid-market" in text:
        return "primært til mindre og mellemstore virksomheder"
    if "mid-market" in text and "enterprise" in text:
        return "primært til mellemstore og større organisationer"
    if "enterprise" in text:
        return "primært til større organisationer"
    if "mid-market" in text:
        return "primært til mellemstore organisationer"
    return "på tværs af virksomhedsstørrelser"


def sector_fit_text(sector_fit: str) -> str:
    lower = sector_fit.lower()
    if "cross-sector" in lower:
        if "critical sectors" in lower:
            return "med fokus på kritiske og regulerede sektorer"
        return "på tværs af brancher"
    labels = []
    if "energy" in lower or "energi" in lower:
        labels.append("energi")
    if "transport" in lower:
        labels.append("transport")
    if "health" in lower or "healthcare" in lower or "sundhed" in lower:
        labels.append("sundhed")
    if "finance" in lower or "financial" in lower or "finans" in lower:
        labels.append("finans")
    if "saas" in lower:
        labels.append("SaaS")
    if "software" in lower or "digital" in lower or "it" in lower:
        labels.append("it og digitale forretninger")
    if not labels:
        return "med fokus på regulerede og forretningskritiske miljøer"
    joined = ", ".join(dict.fromkeys(labels))
    return f"med særlig relevans for {joined}"


def keyword_present(text: str, keywords: list[str]) -> bool:
    lowered = text.lower()
    return any(keyword in lowered for keyword in keywords)


def build_best_for(primary_type: str, seed: str, size_fit: str, sector_fit: str) -> str:
    seed_lower = seed.lower()
    size_text = size_fit_text(size_fit)
    sector_text = sector_fit_text(sector_fit)

    if primary_type == "Legal":
        phrases = ["Juridisk rådgivning om NIS2, cybersikkerhed og regulatoriske krav"]
        if keyword_present(seed_lower, ["contract", "kontrakt", "outsourcing"]):
            phrases.append("med fokus på kontrakter og leverandørforhold")
        if keyword_present(seed_lower, ["privacy", "data", "gdpr", "databeskyttelse"]):
            phrases.append("samt databeskyttelse og informationsforvaltning")
        if keyword_present(seed_lower, ["public sector", "offentlig", "myndighed"]):
            phrases.append("og erfaring fra offentlige eller stærkt regulerede miljøer")
        if keyword_present(seed_lower, ["cer", "cra"]):
            phrases.append("inklusive relaterede EU-regler som CER og CRA")
        phrases.append(size_text)
        phrases.append(sector_text)
        return ", ".join(phrases[:4]).replace(", og", " og")

    if primary_type == "GRC":
        phrases = [
            "GRC-, ISMS- og dokumentationsforløb med fokus på NIS2, politikker, risikostyring og leverandørstyring"
        ]
        if keyword_present(seed_lower, ["iso", "isms"]):
            phrases.append("inklusive standardnære kontrolrammer og governance")
        if keyword_present(seed_lower, ["privacy", "gdpr", "third-party", "supplier"]):
            phrases.append("samt leverandørstyring, privacy eller tredjepartsrisiko")
        if keyword_present(seed_lower, ["identity", "access"]):
            phrases.append("med kobling til identitets- og adgangsgovernance")
        phrases.append(size_text)
        return ", ".join(phrases[:4]).replace(", og", " og")

    if primary_type == "Technical":
        phrases = [
            "Teknisk implementering af sikkerhedskontroller, arkitektur, adgangsstyring, logging og driftsnære forbedringer"
        ]
        if keyword_present(seed_lower, ["identity", "iam", "pam", "adgang"]):
            phrases.append("med fokus på identitet, adgangsstyring og privilegerede konti")
        if keyword_present(seed_lower, ["network", "segment", "firewall"]):
            phrases.append("inklusive netværkssikkerhed, segmentering og hardening")
        if keyword_present(seed_lower, ["backup", "recovery", "continuity"]):
            phrases.append("samt backup, recovery og robust it-drift")
        if keyword_present(seed_lower, ["cloud", "microsoft"]):
            phrases.append("med tydelig cloud- og platformskompetence")
        phrases.append(size_text)
        return ", ".join(phrases[:5]).replace(", og", " og")

    if primary_type == "SOC":
        phrases = [
            "SOC-, logging- og monitoreringskapacitet med fokus på detektion, MDR, incident response og løbende sikkerhedsdrift"
        ]
        if keyword_present(seed_lower, ["24/7", "continuous", "managed", "soc"]):
            phrases.append("til virksomheder der ønsker mere kontinuerlig sikkerhedsdrift")
        if keyword_present(seed_lower, ["cloud", "endpoint", "xdr"]):
            phrases.append("med styrke indenfor endpoint-, cloud- eller XDR-baseret overvågning")
        phrases.append(size_text)
        return ", ".join(phrases[:4]).replace(", og", " og")

    phrases = [
        "Uafhængig assurance, review og modenhedsvurderinger af NIS2-relaterede kontroller"
    ]
    if keyword_present(seed_lower, ["certification", "certific", "assessment"]):
        phrases.append("inklusive certificerings- eller vurderingsforløb")
    if keyword_present(seed_lower, ["supplier", "third-party", "leverandør"]):
        phrases.append("med fokus på dokumentation og leverandørkrav")
    phrases.append(size_text)
    return ", ".join(phrases[:4]).replace(", og", " og")


def build_row_from_template(
    template_row: dict[str, str], entry: dict[str, str], primary_type: str
) -> dict[str, str]:
    row = deepcopy(template_row)
    row["Company"] = entry["name"]
    row["Primary_type"] = primary_type
    row["Website"] = entry["website"]
    row["Qualification_method_note"] = QUALIFICATION_NOTE
    row["Qualification_score_initial"] = str(entry.get("score", "75"))
    row["Size_fit"] = entry.get("size_fit", template_row["Size_fit"])
    row["Sector_fit"] = entry.get("sector_fit", template_row["Sector_fit"])
    row["Price_band"] = entry.get("price_band", template_row["Price_band"])
    row["Secondary_types"] = entry.get("secondary_types", template_row["Secondary_types"])
    row["Adjacent_types"] = entry.get("adjacent_types", template_row["Adjacent_types"])
    row["Best_for"] = build_best_for(
        primary_type,
        entry["best_for"],
        row["Size_fit"],
        row["Sector_fit"],
    )
    row["Recommended_role"] = entry.get(
        "recommended_role", DEFAULT_ROLE_BY_TYPE[primary_type]
    )
    row["Recommended_when"] = entry.get(
        "recommended_when", DEFAULT_ROLE_BY_TYPE[primary_type]
    )
    return row


def translate_existing_row(row: dict[str, str]) -> dict[str, str]:
    updated = deepcopy(row)
    updated["Qualification_method_note"] = QUALIFICATION_NOTE
    updated["Best_for"] = build_best_for(
        updated["Primary_type"],
        updated["Best_for"],
        updated["Size_fit"],
        updated["Sector_fit"],
    )
    updated["Recommended_role"] = DEFAULT_ROLE_BY_TYPE[updated["Primary_type"]]
    updated["Recommended_when"] = DEFAULT_ROLE_BY_TYPE[updated["Primary_type"]]
    return updated


def dedupe_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    seen: set[tuple[str, str]] = set()
    deduped: list[dict[str, str]] = []
    for row in rows:
        key = (row["Primary_type"], row["Company"].strip().lower())
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def rerank_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    by_type: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        by_type[row["Primary_type"]].append(row)

    reranked: list[dict[str, str]] = []
    for vendor_type in TYPE_SORT_ORDER:
        sorted_rows = sorted(
            by_type[vendor_type],
            key=lambda item: (
                -int(float(item["Qualification_score_initial"] or 0)),
                item["Company"].lower(),
            ),
        )
        for index, row in enumerate(sorted_rows, start=1):
            row["Rank_in_type"] = str(index)
            reranked.append(row)
    return reranked


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, rows: list[dict[str, str]]) -> None:
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_xlsx(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Kategori", "Antal profiler"])

    counter = Counter(row["Primary_type"] for row in rows)
    for vendor_type in TYPE_SORT_ORDER:
        summary_sheet.append([vendor_type, counter[vendor_type]])
    summary_sheet.append(["Total", len(rows)])

    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)

    all_sheet = workbook.create_sheet("Vendors_All")
    all_sheet.append(headers)
    for row in rows:
        all_sheet.append([row.get(header, "") for header in headers])
    for cell in all_sheet[1]:
        cell.font = Font(bold=True)

    for vendor_type in TYPE_SORT_ORDER:
        sheet = workbook.create_sheet(vendor_type)
        vendor_rows = [row for row in rows if row["Primary_type"] == vendor_type]
        sheet.append(headers)
        for row in vendor_rows:
            sheet.append([row.get(header, "") for header in headers])
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    for sheet in workbook.worksheets:
        max_widths: dict[int, int] = defaultdict(int)
        for row in sheet.iter_rows(values_only=True):
            for index, value in enumerate(row, start=1):
                width = len(str(value or ""))
                if width > max_widths[index]:
                    max_widths[index] = min(width, 80)
        for index, width in max_widths.items():
            sheet.column_dimensions[get_column_letter(index)].width = width + 2
        sheet.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    rows, headers = load_csv_rows()
    deduped_rows = dedupe_rows(rows)
    template_lookup = {row["Company"]: row for row in deduped_rows}
    existing_keys = {
        (row["Primary_type"], row["Company"].strip().lower()) for row in deduped_rows
    }

    transformed_rows = [translate_existing_row(row) for row in deduped_rows]

    for primary_type, additions in ADDITIONS.items():
        template = template_lookup[TEMPLATE_BY_TYPE[primary_type]]
        for addition in additions:
            addition_key = (primary_type, addition["name"].strip().lower())
            if addition_key in existing_keys:
                continue
            transformed_rows.append(build_row_from_template(template, addition, primary_type))

    reranked_rows = rerank_rows(transformed_rows)

    counts = Counter(row["Primary_type"] for row in reranked_rows)
    for vendor_type, target_count in TARGET_COUNTS.items():
        if counts[vendor_type] != target_count:
            raise SystemExit(
                f"Expected {target_count} rows for {vendor_type}, got {counts[vendor_type]}"
            )

    write_csv(CSV_PATH, headers, reranked_rows)
    write_json(JSON_PATH, reranked_rows)
    write_xlsx(XLSX_PATH, headers, reranked_rows)
    write_xlsx(DOWNLOAD_XLSX_PATH, headers, reranked_rows)

    print("Updated vendor catalog:")
    for vendor_type in TYPE_SORT_ORDER:
        print(f"  {vendor_type}: {counts[vendor_type]}")
    print(f"  Total: {len(reranked_rows)}")
    print(f"Excel written to: {XLSX_PATH}")
    print(f"Excel copied to: {DOWNLOAD_XLSX_PATH}")


if __name__ == "__main__":
    main()
